"""
Procesador de archivos para extraer texto
Soporta: TXT, PDF, DOCX, JSON, CSV, XLSX
"""

import PyPDF2
from docx import Document
import openpyxl
import json
import csv
from typing import Optional, Dict, Any
from io import StringIO, BytesIO

class FileProcessor:
    """Clase para procesar diferentes formatos de archivo"""
    
    SUPPORTED_FORMATS = {
        'txt': 'text/plain',
        'pdf': 'application/pdf',
        'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'json': 'application/json',
        'csv': 'text/csv',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'xls': 'application/vnd.ms-excel'
    }
    
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB enough for 20000word doc
    
    @staticmethod
    def get_file_extension(filename: str) -> str:
        """Obtiene la extensión del archivo"""
        return filename.split('.')[-1].lower()
    
    @staticmethod
    def is_supported(filename: str) -> bool:
        """Verifica si el formato es soportado"""
        ext = FileProcessor.get_file_extension(filename)
        return ext in FileProcessor.SUPPORTED_FORMATS
    
    @staticmethod
    async def process_file(file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Procesa el archivo y extrae el texto
        
        Args:
            file_content: Contenido del archivo en bytes
            filename: Nombre del archivo
            
        Returns:
            Dict con el texto extraído y metadata
        """
        ext = FileProcessor.get_file_extension(filename)
        
        if not FileProcessor.is_supported(filename):
            return {
                'success': False,
                'error': f'Formato no soportado: {ext}',
                'supported_formats': list(FileProcessor.SUPPORTED_FORMATS.keys())
            }
        
        if len(file_content) > FileProcessor.MAX_FILE_SIZE:
            return {
                'success': False,
                'error': f'Archivo demasiado grande. Máximo: {FileProcessor.MAX_FILE_SIZE / 1024 / 1024} MB'
            }
        
        try:
            # Procesar según el tipo de archivo
            if ext == 'txt':
                text = FileProcessor._process_txt(file_content)
            elif ext == 'pdf':
                text = FileProcessor._process_pdf(file_content)
            elif ext == 'docx':
                text = FileProcessor._process_docx(file_content)
            elif ext == 'json':
                text = FileProcessor._process_json(file_content)
            elif ext == 'csv':
                text = FileProcessor._process_csv(file_content)
            elif ext in ['xlsx', 'xls']:
                text = FileProcessor._process_xlsx(file_content)
            else:
                return {
                    'success': False,
                    'error': f'Procesador no implementado para: {ext}'
                }
            
            return {
                'success': True,
                'text': text,
                'filename': filename,
                'format': ext,
                'size_bytes': len(file_content),
                'char_count': len(text)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Error procesando archivo: {str(e)}',
                'filename': filename
            }
    
    @staticmethod
    def _process_txt(content: bytes) -> str:
        """Procesa archivos TXT"""
        return content.decode('utf-8', errors='ignore').strip()
    
    @staticmethod
    def _process_pdf(content: bytes) -> str:
        """Procesa archivos PDF"""
        pdf_file = BytesIO(content)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        
        text_parts = []
        for page in pdf_reader.pages:
            text_parts.append(page.extract_text())
        
        return '\n\n'.join(text_parts).strip()
    
    @staticmethod
    def _process_docx(content: bytes) -> str:
        """Procesa archivos DOCX"""
        docx_file = BytesIO(content)
        doc = Document(docx_file)
        
        text_parts = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)
        
        return '\n\n'.join(text_parts).strip()
    
    @staticmethod
    def _process_json(content: bytes) -> str:
        """Procesa archivos JSON - extrae todos los valores de texto"""
        json_str = content.decode('utf-8', errors='ignore')
        data = json.loads(json_str)
        
        def extract_text(obj):
            """Extrae recursivamente todo el texto del JSON"""
            texts = []
            if isinstance(obj, dict):
                for value in obj.values():
                    texts.extend(extract_text(value))
            elif isinstance(obj, list):
                for item in obj:
                    texts.extend(extract_text(item))
            elif isinstance(obj, str):
                texts.append(obj)
            return texts
        
        all_texts = extract_text(data)
        return '\n'.join(all_texts).strip()
    
    @staticmethod
    def _process_csv(content: bytes) -> str:
        """Procesa archivos CSV - extrae todas las celdas"""
        csv_str = content.decode('utf-8', errors='ignore')
        csv_file = StringIO(csv_str)
        reader = csv.reader(csv_file)
        
        text_parts = []
        for row in reader:
            row_text = ' | '.join(cell.strip() for cell in row if cell.strip())
            if row_text:
                text_parts.append(row_text)
        
        return '\n'.join(text_parts).strip()
    
    @staticmethod
    def _process_xlsx(content: bytes) -> str:
        """Procesa archivos XLSX - extrae todas las celdas"""
        xlsx_file = BytesIO(content)
        workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
        
        text_parts = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"=== {sheet_name} ===")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = ' | '.join(str(cell).strip() for cell in row if cell is not None and str(cell).strip())
                if row_text:
                    text_parts.append(row_text)
        
        return '\n'.join(text_parts).strip()